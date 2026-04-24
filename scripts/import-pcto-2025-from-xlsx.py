import json
import os
import ssl
import sys
import urllib.error
import urllib.parse
import urllib.request
from datetime import date, datetime, time
from zoneinfo import ZoneInfo

from openpyxl import load_workbook


SOURCE_SPREADSHEET_ID = "1Wa9EOEsUzt_ii7cqOXLoaZ8dRlC19VMl3hHJfliGMfw"
SCHOOL_YEAR_LABEL = "2025/2026"
LOCAL_TIMEZONE = ZoneInfo("Europe/Rome")
REGISTRATIONS_SHEET = "Iscritti"
ATTENDANCE_SHEET = "Presenze"

REGISTRATION_HEADERS = {
    "source_code": "Code",
    "duplicate_code": "Dup",
    "registration_submitted_at": "Inserimento",
    "student_first_name": "Nome",
    "student_last_name": "Cognome",
    "display_name": "Name_Surname",
    "school_name": "Scuola",
    "registration_status": "Stato",
    "waiting_list_position": "Attesa",
    "assigned_service_name": "Luogo",
    "attendance_count": "Presenze",
    "class_year": "Classe",
    "class_section": "Sezione",
    "teacher_name": "Professore di riferimento per la PCTO nella tua scuola",
    "student_phone": "Telefono",
    "student_email": "Email",
    "student_address": "Indirizzo",
    "certificate_type": "Tipo",
    "friend_preferences": "Amici",
    "unavailable_days": "Non_disponibile",
    "student_notes": "Note",
    "internal_notes": "Note_nostre",
    "registry_confirmed": "Registro",
    "invitation_sent": "Invito",
    "duplicate_marker": "Doppione",
}

ATTENDANCE_HEADERS = {
    "submitted_at": "Informazioni cronologiche",
    "source_code": "Il tuo codice ID",
    "student_first_name": "Nome",
    "student_last_name": "Cognome",
    "service_name": "Dove stai facendo servizio",
    "service_date": "Data del servizio",
    "check_in_time": "Ora di entrata",
    "check_out_time": "Ora di uscita",
    "notes": "Se hai bisogno di comunicarci qualcosa scrivila pure qui",
}


def normalize_text(value):
    if value is None:
        return ""
    if isinstance(value, float) and value.is_integer():
        value = int(value)
    return " ".join(str(value).replace("\r", " ").split()).strip()


def blank_to_none(value):
    text = normalize_text(value)
    return text or None


def parse_int(value):
    if value in (None, ""):
        return None

    try:
        return int(float(str(value).strip()))
    except ValueError:
        return None


def parse_bool(value):
    text = normalize_text(value).lower()
    if not text:
        return None
    if text in {"true", "vero", "sì", "si", "yes", "1"}:
        return True
    if text in {"false", "falso", "no", "0"}:
        return False
    return None


def parse_certificate_type(value):
    text = normalize_text(value).lower()
    if text == "volontariato":
        return "volontariato"
    return "pcto"


def parse_datetime(value):
    if value in (None, ""):
        return None
    if isinstance(value, datetime):
        dt = value
    elif isinstance(value, date):
        dt = datetime.combine(value, time.min)
    else:
        text = normalize_text(value)
        formats = (
            "%d/%m/%Y %H.%M.%S",
            "%d/%m/%Y %H:%M:%S",
            "%Y-%m-%d %H:%M:%S.%f",
            "%Y-%m-%d %H:%M:%S",
        )
        for fmt in formats:
            try:
                dt = datetime.strptime(text, fmt)
                break
            except ValueError:
                continue
        else:
            return None

    if dt.tzinfo is None:
        dt = dt.replace(tzinfo=LOCAL_TIMEZONE)
    return dt.isoformat()


def parse_date(value):
    if value in (None, ""):
        return None
    if isinstance(value, datetime):
        return value.date().isoformat()
    if isinstance(value, date):
        return value.isoformat()

    text = normalize_text(value)
    formats = ("%d/%m/%Y", "%Y-%m-%d %H:%M:%S", "%Y-%m-%d")
    for fmt in formats:
        try:
            return datetime.strptime(text, fmt).date().isoformat()
        except ValueError:
            continue
    return None


def parse_time(value):
    if value in (None, ""):
        return None
    if isinstance(value, datetime):
        return value.time().replace(microsecond=0).isoformat()
    if isinstance(value, time):
        return value.replace(microsecond=0).isoformat()

    text = normalize_text(value)
    formats = ("%H.%M.%S", "%H:%M:%S", "%H.%M", "%H:%M")
    for fmt in formats:
        try:
            return datetime.strptime(text, fmt).time().isoformat()
        except ValueError:
            continue
    return None


def duration_minutes(check_in, check_out):
    if not check_in or not check_out:
        return None
    start = datetime.strptime(check_in, "%H:%M:%S")
    end = datetime.strptime(check_out, "%H:%M:%S")
    delta = end - start
    minutes = int(delta.total_seconds() // 60)
    return minutes if minutes >= 0 else None


def serialize_raw_value(value):
    if isinstance(value, datetime):
        return value.isoformat()
    if isinstance(value, date):
        return value.isoformat()
    if isinstance(value, time):
        return value.isoformat()
    return value


def row_to_dict(headers, values):
    return {
        header: serialize_raw_value(values[index])
        for index, header in enumerate(headers)
        if header and index < len(values)
    }


def read_sheet_rows(workbook, sheet_name):
    sheet = workbook[sheet_name]
    rows = list(sheet.iter_rows(values_only=True))
    headers = [normalize_text(value) for value in rows[0]]

    parsed_rows = []
    for row_number, values in enumerate(rows[1:], start=2):
        if not any(value not in (None, "") for value in values):
            continue
        raw_data = row_to_dict(headers, values)
        by_header = {
            headers[index]: values[index] if index < len(values) else None
            for index in range(len(headers))
        }
        parsed_rows.append((row_number, by_header, raw_data))

    return parsed_rows


def rest_request(method, url, headers, payload=None):
    body = None
    if payload is not None:
        body = json.dumps(payload).encode("utf-8")

    request = urllib.request.Request(url, data=body, method=method)
    for key, value in headers.items():
        request.add_header(key, value)

    ssl_context = ssl.create_default_context()

    try:
        with urllib.request.urlopen(request, context=ssl_context) as response:
            content = response.read().decode("utf-8")
            return json.loads(content) if content else None
    except urllib.error.HTTPError as error:
        details = error.read().decode("utf-8")
        raise RuntimeError(f"{error.code} {details}") from error


def chunked(items, size=100):
    for index in range(0, len(items), size):
        yield items[index : index + size]


def build_registration_payload(row_number, row, raw_data, school_year_id):
    payload = {
        "school_year_id": school_year_id,
        "source_spreadsheet_id": SOURCE_SPREADSHEET_ID,
        "source_sheet_name": REGISTRATIONS_SHEET,
        "source_row_number": row_number,
        "raw_data": raw_data,
    }

    for field, header in REGISTRATION_HEADERS.items():
        value = row.get(header)
        if field == "registration_submitted_at":
            payload[field] = parse_datetime(value)
        elif field in {"waiting_list_position", "attendance_count"}:
            payload[field] = parse_int(value)
        elif field in {"registry_confirmed", "invitation_sent"}:
            payload[field] = parse_bool(value)
        elif field == "certificate_type":
            payload[field] = parse_certificate_type(value)
        elif field in {"source_code", "student_first_name", "student_last_name"}:
            payload[field] = normalize_text(value)
        else:
            payload[field] = blank_to_none(value)

    return payload


def build_attendance_payload(row_number, row, raw_data, school_year_id, registrations_by_code):
    check_in = parse_time(row.get(ATTENDANCE_HEADERS["check_in_time"]))
    check_out = parse_time(row.get(ATTENDANCE_HEADERS["check_out_time"]))
    source_code = normalize_text(row.get(ATTENDANCE_HEADERS["source_code"]))

    return {
        "school_year_id": school_year_id,
        "student_registration_id": registrations_by_code.get(source_code),
        "source_spreadsheet_id": SOURCE_SPREADSHEET_ID,
        "source_sheet_name": ATTENDANCE_SHEET,
        "source_row_number": row_number,
        "source_code": source_code,
        "submitted_at": parse_datetime(row.get(ATTENDANCE_HEADERS["submitted_at"])),
        "student_first_name": blank_to_none(row.get(ATTENDANCE_HEADERS["student_first_name"])),
        "student_last_name": blank_to_none(row.get(ATTENDANCE_HEADERS["student_last_name"])),
        "service_name": blank_to_none(row.get(ATTENDANCE_HEADERS["service_name"])),
        "service_date": parse_date(row.get(ATTENDANCE_HEADERS["service_date"])),
        "check_in_time": check_in,
        "check_out_time": check_out,
        "duration_minutes": duration_minutes(check_in, check_out),
        "notes": blank_to_none(row.get(ATTENDANCE_HEADERS["notes"])),
        "raw_data": raw_data,
    }


def main():
    if len(sys.argv) < 2:
        print(
            "Usage: python3 scripts/import-pcto-2025-from-xlsx.py <xlsx-path> [--replace]",
            file=sys.stderr,
        )
        sys.exit(1)

    supabase_url = os.environ.get("NEXT_PUBLIC_SUPABASE_URL")
    service_role_key = os.environ.get("SUPABASE_SERVICE_ROLE_KEY")

    if not supabase_url or not service_role_key:
        print("Missing Supabase environment variables.", file=sys.stderr)
        sys.exit(1)

    xlsx_path = os.path.abspath(sys.argv[1])
    replace_existing = "--replace" in sys.argv[2:]
    workbook = load_workbook(xlsx_path, data_only=True, read_only=True)

    headers = {
        "apikey": service_role_key,
        "Authorization": f"Bearer {service_role_key}",
        "Content-Type": "application/json",
    }

    school_year_query = urllib.parse.urlencode(
        {"select": "id,label", "label": f"eq.{SCHOOL_YEAR_LABEL}"}
    )
    school_years = rest_request(
        "GET",
        f"{supabase_url}/rest/v1/school_years?{school_year_query}",
        headers,
    ) or []

    if not school_years:
        print(f"School year {SCHOOL_YEAR_LABEL} not found.", file=sys.stderr)
        sys.exit(1)

    school_year_id = school_years[0]["id"]

    if replace_existing:
        source_query = urllib.parse.urlencode(
            {
                "school_year_id": f"eq.{school_year_id}",
                "source_spreadsheet_id": f"eq.{SOURCE_SPREADSHEET_ID}",
            }
        )
        rest_request(
            "DELETE",
            f"{supabase_url}/rest/v1/pcto_attendance_records?{source_query}",
            {**headers, "Prefer": "return=minimal"},
        )
        rest_request(
            "DELETE",
            f"{supabase_url}/rest/v1/pcto_student_registrations?{source_query}",
            {**headers, "Prefer": "return=minimal"},
        )

    registration_rows = read_sheet_rows(workbook, REGISTRATIONS_SHEET)
    registration_payloads = [
        build_registration_payload(row_number, row, raw_data, school_year_id)
        for row_number, row, raw_data in registration_rows
    ]
    registration_payloads = [
        payload
        for payload in registration_payloads
        if payload["source_code"]
        and payload["student_first_name"]
        and payload["student_last_name"]
    ]

    registration_conflict = urllib.parse.urlencode(
        {"on_conflict": "school_year_id,source_code"}
    )
    for batch in chunked(registration_payloads):
        rest_request(
            "POST",
            f"{supabase_url}/rest/v1/pcto_student_registrations?{registration_conflict}",
            {**headers, "Prefer": "resolution=merge-duplicates"},
            batch,
        )

    registration_select = urllib.parse.urlencode(
        {
            "select": "id,source_code",
            "school_year_id": f"eq.{school_year_id}",
        }
    )
    existing_registrations = rest_request(
        "GET",
        f"{supabase_url}/rest/v1/pcto_student_registrations?{registration_select}",
        headers,
    ) or []
    registrations_by_code = {
        item["source_code"]: item["id"] for item in existing_registrations
    }

    attendance_rows = read_sheet_rows(workbook, ATTENDANCE_SHEET)
    attendance_payloads = [
        build_attendance_payload(
            row_number,
            row,
            raw_data,
            school_year_id,
            registrations_by_code,
        )
        for row_number, row, raw_data in attendance_rows
    ]
    attendance_payloads = [
        payload for payload in attendance_payloads if payload["source_code"]
    ]

    attendance_conflict = urllib.parse.urlencode(
        {"on_conflict": "source_spreadsheet_id,source_sheet_name,source_row_number"}
    )
    for batch in chunked(attendance_payloads):
        rest_request(
            "POST",
            f"{supabase_url}/rest/v1/pcto_attendance_records?{attendance_conflict}",
            {**headers, "Prefer": "resolution=merge-duplicates"},
            batch,
        )

    linked_attendance_count = sum(
        1 for payload in attendance_payloads if payload["student_registration_id"]
    )

    print(f"Imported or updated student registrations: {len(registration_payloads)}")
    print(f"Imported or updated attendance records: {len(attendance_payloads)}")
    print(f"Attendance records linked to registrations: {linked_attendance_count}")


if __name__ == "__main__":
    main()
